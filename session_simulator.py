import tkinter as tk
import secrets
import time
import random
import string
import datetime

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# -----------------------
# Globals
# -----------------------
current_session = None
current_user = None
secure_mode = False
session_time = None

# ✅ NEW (for fixing export only)
current_attack_name = "General"
structured_logs = []

# -----------------------
# Logging (UI unchanged)
# -----------------------
def log(msg, color="white"):
    global structured_logs

    timestamp = time.strftime("%H:%M:%S")

    # ✅ Structured logging (NEW)
    status = ""
    if "SUCCESS" in msg:
        status = "SUCCESS"
    elif "BLOCKED" in msg:
        status = "BLOCKED"

    structured_logs.append({
        "time": timestamp,
        "attack": current_attack_name,
        "action": msg.replace("✔", "").replace("✖", "").strip(),
        "status": status
    })

    # ---- EXISTING UI LOG (UNCHANGED) ----
    log_box.config(state=tk.NORMAL)

    start = log_box.index("end-1c")
    log_box.insert(tk.END, f"[{timestamp}] {msg}\n")
    end = log_box.index("end-1c")

    log_box.tag_add(color, start, end)

    log_box.tag_config("red", foreground="#ff4d4d")
    log_box.tag_config("green", foreground="#00ff9f")
    log_box.tag_config("yellow", foreground="#f2cc60")
    log_box.tag_config("white", foreground="white")

    log_box.config(state=tk.DISABLED)
    log_box.see(tk.END)

def clear_logs():
    log_box.config(state=tk.NORMAL)
    log_box.delete(1.0, tk.END)
    log_box.config(state=tk.DISABLED)

# -----------------------
# Excel Export (FIXED)
# -----------------------
def export_logs():
    if not structured_logs:
        log("No logs to export", "red")
        return

    filename = f"session_logs_{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Session Logs"

    headers = ["Time", "Attack", "Action", "Status"]
    ws.append(headers)

    for entry in structured_logs:
        ws.append([
            entry["time"],
            entry["attack"],
            entry["action"],
            entry["status"]
        ])

    # Auto width
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 5

    # Color status
    for row in ws.iter_rows(min_row=2):
        status_cell = row[3]
        if status_cell.value == "SUCCESS":
            status_cell.font = Font(color="00AA00", bold=True)
        elif status_cell.value == "BLOCKED":
            status_cell.font = Font(color="FF0000", bold=True)

    wb.save(filename)
    log(f"Logs exported → {filename}", "green")

# -----------------------
# Mode Toggle
# -----------------------
def toggle_mode():
    global secure_mode
    secure_mode = not secure_mode

    if secure_mode:
        mode_label.config(text="SECURE MODE", fg="green")
        log("Defense Enabled", "green")
    else:
        mode_label.config(text="VULNERABLE MODE", fg="red")
        log("System Vulnerable", "red")

# -----------------------
# Login
# -----------------------
def login():
    global current_session, current_user, session_time

    username = username_entry.get()
    if not username:
        log("Enter username", "red")
        return

    current_user = username
    session_time = time.time()

    if secure_mode:
        current_session = secrets.token_hex(16)
    else:
        current_session = username + "_session"

    user_label.config(text=f"User: {current_user}")
    session_label.config(text=f"Session: {current_session}")

    log(f"User '{username}' logged in")
    log(f"Session Created: {current_session}")

# -----------------------
# Clear Username
# -----------------------
def clear_username():
    global current_user, current_session
    username_entry.delete(0, tk.END)

    current_user = None
    current_session = None

    user_label.config(text="User: None")
    session_label.config(text="Session: None")

    log("Username cleared", "yellow")

# -----------------------
# ATTACKS (ONLY 1 LINE ADDED EACH)
# -----------------------
def prediction_attack():
    global current_attack_name
    current_attack_name = "Prediction Attack"

    log("\n--- Prediction Attack ---")
    guess = current_user + "_session"
    log(f"Trying: {guess}")

    if guess == current_session:
        log("✔ SUCCESS: Session Hijacked", "green")
    else:
        log("✖ BLOCKED: Secure Session", "red")

def sniffing_attack():
    global current_attack_name
    current_attack_name = "Session Sniffing Attack"

    log("\n--- Session Sniffing Attack ---")

    if secure_mode:
        log("✖ BLOCKED: Encrypted Traffic", "red")
    else:
        log(f"Captured session: {current_session}")
        log("✔ SUCCESS: Session Stolen", "green")

def sidejacking_attack():
    global current_attack_name
    current_attack_name = "Session Sidejacking Attack"

    log("\n--- Session Sidejacking Attack ---")

    if secure_mode:
        log("✖ BLOCKED: IP mismatch detected", "red")
    else:
        log("✔ SUCCESS: Unauthorized access granted", "green")

def brute_force_attack():
    global current_attack_name
    current_attack_name = "Session ID Brute Force"

    log("\n--- Session ID Brute Force ---")

    if secure_mode:
        for _ in range(5):
            guess = ''.join(random.choices(string.hexdigits, k=8))
            log(f"Guess: {guess}")
        log("✖ BLOCKED: Strong session ID", "red")
    else:
        guess = current_user + "_session"
        log(f"Guess: {guess}")
        log("✔ SUCCESS: Weak session guessed", "green")

def timeout_attack():
    global current_attack_name
    current_attack_name = "Session Timeout Attack"

    log("\n--- Session Timeout Attack ---")

    elapsed = time.time() - session_time
    log(f"Session age: {int(elapsed)} sec")

    if secure_mode:
        log("✖ BLOCKED: Session protected", "red")
    else:
        log("✔ SUCCESS: Session never expires", "green")

# -----------------------
# Simulation
# -----------------------
def run_simulation():
    if not current_user:
        log("Login first", "red")
        return

    structured_logs.clear()
    clear_logs()

    log("=== STARTING CYBER ATTACK SIMULATION ===\n", "yellow")

    root.after(1000, prediction_attack)
    root.after(2500, sniffing_attack)
    root.after(4000, sidejacking_attack)
    root.after(5500, brute_force_attack)
    root.after(7000, timeout_attack)
    root.after(9000, lambda: log("\n=== SIMULATION COMPLETE ===", "yellow"))

# -----------------------
# GUI (UNCHANGED)
# -----------------------
root = tk.Tk()
root.title("Cyber Range Session Hijacking Simulator")
root.geometry("800x580")
root.configure(bg="#0d1117")

title = tk.Label(root, text="CYBER RANGE SIMULATION",
                 font=("Consolas", 18), bg="#0d1117", fg="#58a6ff")
title.pack(pady=10)

mode_label = tk.Label(root, text="VULNERABLE MODE", fg="red", bg="#0d1117")
mode_label.pack()

tk.Button(root, text="Toggle Mode", command=toggle_mode).pack(pady=5)

frame = tk.Frame(root, bg="#0d1117")
frame.pack()

username_entry = tk.Entry(frame, width=25)
username_entry.grid(row=0, column=0, padx=5)

tk.Button(frame, text="Login", command=login).grid(row=0, column=1, padx=5)
tk.Button(frame, text="Clear", command=clear_username,
          bg="#da3633", fg="white").grid(row=0, column=2, padx=5)

user_label = tk.Label(root, text="User: None", bg="#0d1117", fg="white")
user_label.pack()

session_label = tk.Label(root, text="Session: None", bg="#0d1117", fg="white")
session_label.pack()

tk.Button(root, text="Run Full Attack Simulation",
          command=run_simulation,
          bg="#238636", fg="white").pack(pady=10)

tk.Button(root, text="Export Logs (Excel)",
          command=export_logs,
          bg="#1f6feb", fg="white").pack(pady=5)

log_box = tk.Text(root, height=15, width=90, bg="#010409", fg="white")
log_box.pack()
log_box.config(state=tk.DISABLED)

root.mainloop()